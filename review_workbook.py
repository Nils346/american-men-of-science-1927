"""Build review_workbook.xlsx: the hand-verification view of the extraction.

One row per scientist, columns in the order the information appears in a
printed entry, so the eye can move left-to-right down the page while checking:

    name | star | address | field | birth | degrees | career | societies | research

The header row and the page/name columns stay frozen while scrolling. A final
"QA flags" column carries every finding qa_check.py raised for that person,
with the row tinted red (error) or amber (warning) so problem entries can be
filtered or spotted at a glance.

Runs automatically after every panel build; standalone:

    python review_workbook.py
"""
from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RAW_PROFILES_JSON = "scientists_raw.json"
FINDINGS_CSV = "qa_findings.csv"
WORKBOOK_XLSX = "review_workbook.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
STRIPE_FILL = PatternFill("solid", fgColor="F2F2F2")
ERROR_FILL = PatternFill("solid", fgColor="F8CBAD")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_EDGE = Border(bottom=Side(style="hair", color="BFBFBF"))

# (title, width, wrap)
COLUMNS = [
    ("Page", 7, False),
    ("Full name", 30, False),
    ("*", 4, False),
    ("Titles", 10, False),
    ("Mailing address", 34, True),
    ("City", 14, False),
    ("State", 8, False),
    ("Country", 10, False),
    ("Field", 20, True),
    ("Birth place", 22, True),
    ("Birth date", 15, False),
    ("Birth year", 10, False),
    ("Degrees", 34, True),
    ("Career positions", 52, True),
    ("Minor positions", 36, True),
    ("Societies", 30, True),
    ("Research (accomplished)", 40, True),
    ("Research (in progress)", 30, True),
    ("QA flags", 34, True),
]


def _page_of(p: dict):
    return p.get("_source_page") or p.get("source_pdf_page")


def _year(y) -> str:
    return str(y) if y else "?"


def _degrees(p: dict) -> str:
    lines = []
    for d in p.get("education") or []:
        inst = d.get("institution") or "?"
        lines.append("%s, %s, %s" % (d.get("degree_type") or "?", inst, _year(d.get("year"))))
    return "\n".join(lines)


def _career(p: dict) -> str:
    lines = []
    for e in p.get("employment") or []:
        span = "%s-%s" % (_year(e.get("start_year")),
                          "" if e.get("is_current_position") and not e.get("end_year")
                          else _year(e.get("end_year")))
        org = (", " + e["institution_org"]) if e.get("institution_org") else ""
        cur = "  [current 1927]" if e.get("is_current_position") else ""
        lines.append("%s%s, %s%s" % (e.get("position_title") or "?", org, span, cur))
    return "\n".join(lines)


def _minor(p: dict) -> str:
    lines = []
    for e in p.get("minor_positions") or []:
        org = (", " + e["institution_org"]) if e.get("institution_org") else ""
        span = ""
        if e.get("start_year") or e.get("end_year"):
            span = ", %s-%s" % (_year(e.get("start_year")), _year(e.get("end_year")))
        lines.append("%s%s%s" % (e.get("position_title") or "?", org, span))
    return "\n".join(lines)


def _load_findings(path: Path) -> dict[tuple[int, str], list[str]]:
    """(page, surname-prefix of scientist field) -> finding summaries."""
    found: dict[tuple[int, str], list[str]] = defaultdict(list)
    if not path.exists():
        return found
    with path.open(encoding="utf-8") as f:
        for row in csv.DictReader(f, delimiter=";"):
            key = (int(row["page"]), row["scientist"])
            found[key].append("%s: %s%s" % (row["severity"].upper(), row["code"],
                                            " -- " + row["detail"] if row["detail"] else ""))
    return found


def _flags_for(profile: dict, page: int,
               findings: dict[tuple[int, str], list[str]]) -> list[str]:
    """Findings whose 'scientist' field names this profile (or its surname)."""
    name = (profile.get("full_name") or "").strip()
    surname = name.split(",", 1)[0].strip()
    out = []
    for (fpage, fname), msgs in findings.items():
        if fpage != page:
            continue
        if fname and (fname == name or fname == surname or name.startswith(fname)):
            out.extend(msgs)
    return out


def write_workbook(profiles: list[dict], base_dir: Path,
                   out_path: Path | None = None) -> Path:
    out_path = out_path or base_dir / WORKBOOK_XLSX
    findings = _load_findings(base_dir / FINDINGS_CSV)

    wb = Workbook()
    ws = wb.active
    ws.title = "Scientists"

    for col, (title, width, _wrap) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22

    # Stable sort by page only: within a page the model's order IS the print
    # order, which is what the eye follows when verifying against the book.
    profiles = sorted(profiles, key=lambda p: _page_of(p) or 0)
    for r, p in enumerate(profiles, start=2):
        page = _page_of(p)
        flags = _flags_for(p, page, findings) if page else []
        values = [
            page,
            p.get("full_name"),
            "*" if p.get("star_status") else "",
            p.get("titles"),
            p.get("mailing_address"),
            p.get("mailing_city"),
            p.get("mailing_state"),
            p.get("mailing_country"),
            p.get("department"),
            p.get("birth_place"),
            p.get("birth_date"),
            p.get("birth_year"),
            _degrees(p),
            _career(p),
            _minor(p),
            "\n".join(p.get("societies") or []),
            p.get("research_accomplished"),
            p.get("research_in_progress"),
            "\n".join(flags),
        ]
        stripe = STRIPE_FILL if r % 2 == 0 else None
        row_fill = stripe
        if any(f.startswith("ERROR") for f in flags):
            row_fill = ERROR_FILL
        elif flags:
            row_fill = WARN_FILL

        for col, ((_t, _w, wrap), value) in enumerate(zip(COLUMNS, values), start=1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=wrap,
                                       horizontal="center" if col in (1, 3, 12) else "general")
            cell.border = THIN_EDGE
            if row_fill is not None:
                cell.fill = row_fill
        ws.cell(row=r, column=2).font = Font(bold=True)

    ws.freeze_panes = "C2"                      # header row + page & name columns
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(COLUMNS)), len(profiles) + 1)

    # Second sheet: the raw QA findings, one per row, for filtering by code.
    qa = wb.create_sheet("QA findings")
    qa_cols = ["Severity", "Code", "Page", "Scientist", "Detail"]
    for col, title in enumerate(qa_cols, start=1):
        cell = qa.cell(row=1, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for w, letter in zip((10, 26, 7, 30, 90), "ABCDE"):
        qa.column_dimensions[letter].width = w
    r = 2
    for (page, name), msgs in sorted(findings.items()):
        for msg in msgs:
            severity, rest = msg.split(": ", 1)
            code, _, detail = rest.partition(" -- ")
            qa.cell(row=r, column=1, value=severity)
            qa.cell(row=r, column=2, value=code)
            qa.cell(row=r, column=3, value=page)
            qa.cell(row=r, column=4, value=name)
            qa.cell(row=r, column=5, value=detail)
            fill = ERROR_FILL if severity == "ERROR" else WARN_FILL
            for c in range(1, 6):
                qa.cell(row=r, column=c).fill = fill
            r += 1
    qa.freeze_panes = "A2"
    qa.auto_filter.ref = "A1:E%d" % max(2, r - 1)

    wb.save(out_path)
    return out_path


def main() -> None:
    base = Path(__file__).resolve().parent
    raw = base / RAW_PROFILES_JSON
    if not raw.exists():
        raise SystemExit("No %s -- run 'python extract_panel.py --panel-only' first."
                         % RAW_PROFILES_JSON)
    profiles = json.loads(raw.read_text(encoding="utf-8"))
    path = write_workbook(profiles, base)
    print("Wrote %s (%d scientists)" % (path.name, len(profiles)))


if __name__ == "__main__":
    main()
