"""
build_release.py
================
Assemble the publishable dataset in `release/` -- the version that goes online
or to other researchers. No QA flags, no working-file leftovers.

    python build_release.py

Reads the working CSVs plus institution_locations.csv, sorts every table in
book order (PDF page, then surname, then year), tags each institution location
as mailing address / AI / hand, and writes both analysis CSVs and formatted
Excel workbooks.

Outputs in release/:
    scientists_1927.csv / .xlsx
    career_events_1927.csv / .xlsx
    scientist_year_panel_1927.csv / .xlsx
    institution_locations_1927.csv / .xlsx
    codebook.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path

import pandas as pd
import xlsxwriter

from merge_institution_locations import load_location_lookup, source_from_notes

BASE = Path(__file__).resolve().parent
RELEASE = BASE / "release"

PANEL_IN, PANEL_OUT = "scientist_mobility_panel.csv", "scientist_year_panel_1927"
EVENTS_IN, EVENTS_OUT = "scientist_events_long.csv", "career_events_1927"
SUMMARY_IN, SUMMARY_OUT = "scientist_summary.csv", "scientists_1927"
LOCATIONS_IN, LOCATIONS_OUT = "institution_locations.csv", "institution_locations_1927"
CODEBOOK = "codebook.xlsx"

NAVY = "#1F3864"
NAVY_SOFT = "#2E5A88"
WHITE = "#FFFFFF"
ZEBRA = "#F4F7FB"
SOURCE_FILL = {
    "mailing address": "#E2EFDA",
    "AI": "#DDEBF7",
    "hand": "#FFF2CC",
}

# Display labels used only in Excel. CSVs keep the machine column names.
LABELS = {
    "source_pdf_page": "Page",
    "scientist_name": "Scientist",
    "last_name": "Surname",
    "first_name": "Given names",
    "title": "Title",
    "star_status": "Starred",
    "primary_department": "Field",
    "year": "Year",
    "age": "Age",
    "activity_confirmed": "Activity confirmed",
    "is_current_1927_role": "Current 1927 role",
    "record_type": "Record type",
    "start_year": "Start year",
    "end_year": "End year",
    "institution_organization": "Institution",
    "role_or_degree": "Role or degree",
    "institution_city": "Institution city",
    "institution_state_region": "Institution state",
    "institution_country": "Institution country",
    "institution_location_source": "Location source",
    "birth_year": "Birth year",
    "birth_date": "Birth date",
    "birth_city": "Birth city",
    "birth_state": "Birth state",
    "birth_country": "Birth country",
    "mailing_city": "Mailing city",
    "mailing_state": "Mailing state",
    "mailing_country": "Mailing country",
    "research": "Research",
    "n_degrees": "Degrees",
    "n_study_spells": "Study spells",
    "n_positions": "Positions",
    "n_parallel_positions": "Minor positions",
    "societies": "Societies",
    "institution": "Institution",
    "n_events": "Events",
    "city": "City",
    "state_region": "State / region",
    "country": "Country",
    "source": "Source",
    "notes": "Notes",
}

PANEL_FRONT = [
    "source_pdf_page", "scientist_name", "last_name", "first_name", "title",
    "star_status", "year", "age", "activity_confirmed",
]
EVENTS_ORDER = [
    "source_pdf_page", "scientist_name", "last_name", "first_name", "title",
    "star_status", "record_type", "start_year", "end_year", "role_or_degree",
    "institution_organization", "institution_city", "institution_state_region",
    "institution_country", "institution_location_source",
    "birth_year", "birth_date", "birth_city", "birth_state", "birth_country",
    "primary_department", "mailing_city", "mailing_state", "mailing_country",
]
SUMMARY_ORDER = [
    "source_pdf_page", "scientist_name", "last_name", "first_name", "title",
    "star_status", "primary_department",
    "birth_year", "birth_date", "birth_city", "birth_state", "birth_country",
    "mailing_city", "mailing_state", "mailing_country",
    "n_degrees", "n_study_spells", "n_positions", "n_parallel_positions",
    "societies", "research",
]
LOCATION_ORDER = [
    "institution", "n_events", "city", "state_region", "country", "source",
]

PANEL_COLS = [
    ("Page / identity", "PDF page of the printed entry, then the scientist's name. "
     "Rows follow the book: page 14, then 15, …; within a page, alphabetical as printed."),
    ("year", "Calendar year; one row per scientist per year from birth through 1927."),
    ("age", "Age in this year (year − birth year); blank without a birth year."),
    ("activity_confirmed", "1 if the directory confirms any dated activity this year; never interpolated."),
    ("degree_earned_N / degree_institution_N", "Degree(s) conferred this year and where; "
     "'study' marks a degreeless study spell. Numbered slots hold concurrent records."),
    ("position_N / institution_N", "Primary career position(s) held this year. "
     "The current-1927 role occupies slot 1 while active."),
    ("is_current_1927_role", "1 if the position in slot 1 is the role held at printing (1927)."),
    ("parallel_position_N / parallel_institution_N",
     "Minor/parallel roles this year (fellowships, military, editorial, summer posts)."),
    ("birth_year / birth_date / birth_city / birth_state / birth_country",
     "As printed in the entry; blank when the book prints none (never inferred)."),
    ("star_status", "1 if starred (top-1000 scientist, asterisk before the subject)."),
    ("primary_department", "Field of investigation printed in italics."),
    ("mailing_city / mailing_state / mailing_country", "Parsed 1927 mailing-address geography."),
    ("research", "Research subjects, ' | '-separated (accomplished + in progress)."),
]

EVENTS_COLS = [
    ("record_type", "Education, Employment, or MinorPosition."),
    ("start_year / end_year", "Explicit printed span, not expanded. Equal years = a single-year stay; "
     "blank end on a current role = held through 1927."),
    ("institution_organization", "Institution/employer as printed (abbreviations kept)."),
    ("role_or_degree", "Degree letters, 'study' for degreeless spells, or the position title."),
    ("is_current_1927_role", "1 if the role was held at printing."),
    ("institution_city / institution_state_region / institution_country",
     "From the institution-locations bridge; blank where not coded."),
    ("institution_location_source",
     "'mailing address' = taken from a scientist's printed 1927 address that names this "
     "institution; 'AI' = drafted by a classify-then-locate model and not yet hand-verified; "
     "'hand' = typed in by a person. Blank when the institution has no location."),
    ("(identity & entry columns)",
     "Same scientist identity, birth, star, department, mailing and provenance columns as the panel."),
]

SUMMARY_COLS = [
    ("n_degrees", "Degrees with degree letters."),
    ("n_study_spells", "Education records without a degree (study stays)."),
    ("n_positions / n_parallel_positions", "Primary and minor career records."),
    ("societies", "Society memberships as printed, ' | '-separated."),
    ("research", "Research subjects, ' | '-separated (accomplished + in progress)."),
    ("(identity & entry columns)",
     "Same identity, birth, star, department, mailing and provenance columns as the panel."),
]

LOCATION_COLS = [
    ("institution", "Institution string exactly as it appears in the career data (join key)."),
    ("n_events", "Number of career events carrying this string (coverage weight)."),
    ("city / state_region / country", "Assigned location. Blank = not coded, never guessed."),
    ("source", "'mailing address' from the book's own printed addresses; 'AI' from a "
     "classify-then-locate draft (unverified); 'hand' from a person. Dual-city and "
     "non-place strings (societies, journals, 'Berlin and Vienna') stay blank on purpose."),
]


def _cell(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if hasattr(v, "item"):
        try:
            return v.item()
        except (ValueError, AttributeError):
            pass
    return v


def _reorder(df: pd.DataFrame, front: list[str]) -> pd.DataFrame:
    seen = [c for c in front if c in df.columns]
    rest = [c for c in df.columns if c not in seen]
    return df.reindex(columns=seen + rest)


def _book_sort(df: pd.DataFrame, extra: list[str]) -> pd.DataFrame:
    keys = [c for c in ["source_pdf_page", "last_name", "first_name", *extra] if c in df.columns]
    out = df.copy()
    if "source_pdf_page" in out.columns:
        out["source_pdf_page"] = pd.to_numeric(out["source_pdf_page"], errors="coerce")
    for col in extra:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce")
    return out.sort_values(keys, na_position="last", kind="stable").reset_index(drop=True)


def _header_fmt(wb, wrap=False):
    return wb.add_format({
        "bold": True, "font_color": WHITE, "bg_color": NAVY, "font_size": 11,
        "font_name": "Calibri", "align": "center", "valign": "vcenter",
        "text_wrap": wrap, "border": 0,
    })


def _cell_fmt(wb, wrap=False, zebra=False, fill=None, center=False, bold=False,
              top=None):
    spec = {
        "font_name": "Calibri", "font_size": 10, "valign": "vcenter",
        "text_wrap": wrap,
        "bg_color": fill or (ZEBRA if zebra else WHITE),
    }
    if center:
        spec["align"] = "center"
    if bold:
        spec["bold"] = True
    if top:
        spec["top"] = top
        spec["top_color"] = NAVY
    return wb.add_format(spec)


def _col_width(name: str, series: pd.Series) -> float:
    label = LABELS.get(name, name.replace("_", " "))
    sample = series.dropna().astype(str).head(400)
    longest = max([len(label)] + [len(s) for s in sample], default=len(label))
    if name in ("research", "societies", "institution", "institution_organization"):
        return min(max(22, longest * 0.85), 48)
    if name.endswith("_page") or name in ("year", "age", "star_status",
                                          "activity_confirmed", "is_current_1927_role",
                                          "n_degrees", "n_study_spells", "n_positions",
                                          "n_parallel_positions", "n_events"):
        return max(8, min(14, longest + 2))
    return min(max(10, longest + 1), 28)


def _write_sheet(wb, sheet_name: str, df: pd.DataFrame, *,
                 freeze_cols: int = 2, wrap: set[str] | None = None,
                 stripe: bool = False, source_col: str | None = None,
                 constant_memory: bool = False, progress: str | None = None,
                 group_cols: list[str] | None = None) -> None:
    wrap = wrap or set()
    group_cols = [c for c in (group_cols or []) if c in df.columns]
    ws = wb.add_worksheet(sheet_name[:31])
    header = _header_fmt(wb, wrap=True)

    cols = list(df.columns)
    n_rows, n_cols = len(df), len(cols)
    group_idx = [cols.index(c) for c in group_cols]
    ws.freeze_panes(1, freeze_cols)
    ws.set_row(0, 28)
    ws.set_default_row(18)
    if n_rows:
        ws.autofilter(0, 0, n_rows, n_cols - 1)
    ws.set_tab_color(NAVY)
    ws.hide_gridlines(2)

    center_names = {
        "source_pdf_page", "year", "age", "star_status", "activity_confirmed",
        "is_current_1927_role", "n_degrees", "n_study_spells", "n_positions",
        "n_parallel_positions", "n_events", "start_year", "end_year",
        "birth_year", "source", "institution_location_source",
    }

    for c, name in enumerate(cols):
        ws.write(0, c, LABELS.get(name, name.replace("_", " ")), header)
        ws.set_column(c, c, _col_width(name, df[name]))

    records = df.itertuples(index=False, name=None)

    # Large sheets: one body format + write_row (42M per-cell calls would take hours).
    # A thick navy top border marks the first row of each new scientist.
    if constant_memory:
        body = _cell_fmt(wb)
        body_break = _cell_fmt(wb, top=5)
        prev = None
        for r, row in enumerate(records, start=1):
            if progress and r % 50_000 == 0:
                print("    %s %s / %s rows" % (progress, f"{r:,}", f"{n_rows:,}"))
            vals = ["" if (x := _cell(v)) is None else x for v in row]
            key = tuple(row[i] for i in group_idx) if group_idx else None
            new_person = bool(group_idx) and key != prev and r > 1
            if new_person:
                ws.set_row(r, 20)
            ws.write_row(r, 0, vals, body_break if new_person else body)
            prev = key
        return

    even = _cell_fmt(wb)
    odd = _cell_fmt(wb, zebra=True)
    even_wrap = _cell_fmt(wb, wrap=True)
    odd_wrap = _cell_fmt(wb, wrap=True, zebra=True)
    even_center = _cell_fmt(wb, center=True)
    odd_center = _cell_fmt(wb, center=True, zebra=True)
    name_even = _cell_fmt(wb, bold=True)
    name_odd = _cell_fmt(wb, bold=True, zebra=True)
    source_fmts = {
        key: wb.add_format({
            "font_name": "Calibri", "font_size": 10, "valign": "vcenter",
            "align": "center", "bg_color": fill,
        })
        for key, fill in SOURCE_FILL.items()
    }

    for r, row in enumerate(records, start=1):
        zebra = stripe and (r % 2 == 0)
        for c, (name, raw) in enumerate(zip(cols, row)):
            val = _cell(raw)
            if source_col and name == source_col and val in source_fmts:
                fmt = source_fmts[val]
            elif name in ("scientist_name", "institution"):
                fmt = name_odd if zebra else name_even
            elif name in wrap:
                fmt = odd_wrap if zebra else even_wrap
            elif name in center_names:
                fmt = odd_center if zebra else even_center
            else:
                fmt = odd if zebra else even
            if val is None:
                ws.write_blank(r, c, None, fmt)
            else:
                ws.write(r, c, val, fmt)


def _write_readme_sheet(wb, title: str, lines: list[tuple[str, str]]) -> None:
    ws = wb.add_worksheet("Read me")
    ws.set_tab_color(NAVY_SOFT)
    ws.hide_gridlines(2)
    ws.set_column(0, 0, 28)
    ws.set_column(1, 1, 100)
    title_fmt = wb.add_format({
        "bold": True, "font_size": 16, "font_color": NAVY, "font_name": "Calibri",
        "valign": "vcenter",
    })
    key_fmt = wb.add_format({
        "bold": True, "font_name": "Calibri", "font_size": 11, "font_color": NAVY,
        "valign": "top",
    })
    val_fmt = wb.add_format({
        "font_name": "Calibri", "font_size": 11, "text_wrap": True, "valign": "top",
    })
    ws.set_row(0, 28)
    ws.write(0, 0, title, title_fmt)
    ws.write(0, 1, "", title_fmt)
    for r, (k, v) in enumerate(lines, start=2):
        ws.set_row(r, 36 if len(str(v)) > 80 else 20)
        ws.write(r, 0, k, key_fmt)
        ws.write(r, 1, v, val_fmt)
    ws.freeze_panes(2, 0)


def _xlsx(path: Path, constant_memory: bool = False) -> xlsxwriter.Workbook:
    return xlsxwriter.Workbook(str(path), {
        "constant_memory": constant_memory,
        "strings_to_urls": False,
        "nan_inf_to_errors": True,
        "default_format_properties": {"font_name": "Calibri", "font_size": 10},
    })


def write_scientists_xlsx(df: pd.DataFrame, stats: dict) -> None:
    path = RELEASE / (SUMMARY_OUT + ".xlsx")
    print("  writing %s ..." % path.name)
    wb = _xlsx(path)
    _write_readme_sheet(wb, "American Men of Science, 1927 — scientists", [
        ("What this file is",
         "One row per scientist, in the order the directory prints them "
         "(PDF page, then surname). Freeze the Page and Scientist columns and "
         "scroll: you are walking down the book."),
        ("Scientists", "{:,}".format(stats["n_scientists"])),
        ("Starred (top-1000)", "{:,}".format(stats["n_starred"])),
        ("How to read a row",
         "Page → name → star → field → birth → 1927 mailing address → "
         "counts of degrees and jobs → societies → research. "
         "Blank means the book did not print it."),
        ("Starred", "1 = asterisk printed before the italic field (Cattell's top 1,000)."),
        ("Also in this release",
         "career_events_1927.xlsx (dated spells), "
         "scientist_year_panel_1927.xlsx (one row per scientist-year), "
         "institution_locations_1927.xlsx (place bridge, with source), "
         "codebook.xlsx (every column). CSVs sit next to the Excel files."),
    ])
    _write_sheet(wb, "Scientists", df, freeze_cols=2, stripe=True,
                 wrap={"societies", "research", "primary_department"})
    wb.close()


def write_events_xlsx(df: pd.DataFrame, stats: dict) -> None:
    path = RELEASE / (EVENTS_OUT + ".xlsx")
    print("  writing %s (%s rows) ..." % (path.name, f"{len(df):,}"))
    wb = _xlsx(path, constant_memory=True)
    _write_readme_sheet(wb, "American Men of Science, 1927 — career events", [
        ("What this file is",
         "One row per dated degree, job, or minor post, in book order. "
         "Year spans are as printed — not expanded into the panel."),
        ("Events", "{:,}".format(stats["n_events"])),
        ("With a location", stats["inst_coverage"] + " of events that name an institution"),
        ("Location source",
         "mailing address = taken from a printed 1927 address that names the institution. "
         "AI = classify-then-locate draft, not yet hand-verified. "
         "hand = typed in by a person. Blank = not coded, never guessed."),
        ("Scientist breaks",
         "A thick navy line marks where one scientist's events end and the next begin."),
    ])
    _write_sheet(wb, "Career events", df, freeze_cols=2,
                 wrap={"institution_organization", "role_or_degree"},
                 source_col="institution_location_source",
                 constant_memory=True, progress="events",
                 group_cols=["source_pdf_page", "scientist_name"])
    wb.close()


def write_panel_xlsx(df: pd.DataFrame, stats: dict) -> None:
    path = RELEASE / (PANEL_OUT + ".xlsx")
    print("  writing %s (%s rows) ..." % (path.name, f"{len(df):,}"))
    wb = _xlsx(path, constant_memory=True)
    _write_readme_sheet(wb, "American Men of Science, 1927 — scientist-year panel", [
        ("What this file is",
         "The main analysis file. One row per scientist per calendar year from birth "
         "through 1927, in book order. Scroll down a person and the years run like "
         "the printed career; freeze Page + Scientist so the name stays in view."),
        ("Rows", "{:,}".format(stats["n_panel"])),
        ("Scientists", "{:,}".format(stats["n_scientists"])),
        ("Activity confirmed",
         "1 only when the directory itself dates a degree or a spell in that year. "
         "Gaps are left blank — never interpolated."),
        ("Scientist breaks",
         "A thick navy line across the sheet marks where one scientist ends and "
         "the next begins. Freeze Page + Scientist, then scroll: each block is "
         "one person's years."),
    ])
    wrap = {c for c in df.columns if "research" in c}
    _write_sheet(wb, "Scientist-year panel", df, freeze_cols=2, wrap=wrap,
                 constant_memory=True, progress="panel",
                 group_cols=["source_pdf_page", "scientist_name"])
    wb.close()


def write_locations_xlsx(located: pd.DataFrame, blank: pd.DataFrame, stats: dict) -> None:
    path = RELEASE / (LOCATIONS_OUT + ".xlsx")
    print("  writing %s ..." % path.name)
    wb = _xlsx(path)
    _write_readme_sheet(wb, "American Men of Science, 1927 — institution locations", [
        ("What this file is",
         "The place bridge. Join `Institution` to any institution column in the "
         "events or panel files. Only exact string matches."),
        ("Located strings", "{:,} of {:,} distinct institution strings".format(
            stats["n_inst_coded"], stats["n_inst"])),
        ("Event coverage", stats["inst_coverage"]),
        ("mailing address",
         "{:,} strings / {:,} events — taken from a scientist's printed 1927 "
         "mailing address that names this institution (majority vote, abbreviation-aware). "
         "Residence-only segments never count.".format(
             stats["n_mail_rows"], stats["n_mail_events"])),
        ("AI",
         "{:,} strings / {:,} events — drafted by a classify-then-locate model. "
         "The model first decides whether the string is a locatable 1927 place; "
         "ambiguous tours ('Berlin and Vienna') and non-places (societies, journals) "
         "stay blank. These rows are not yet hand-verified.".format(
             stats["n_ai_rows"], stats["n_ai_events"])),
        ("hand",
         "{:,} strings / {:,} events — typed in by a person.".format(
             stats["n_hand_rows"], stats["n_hand_events"])),
        ("Not located",
         "{:,} strings left blank on purpose (not a place, ambiguous, or not yet coded).".format(
             stats["n_blank"])),
        ("Colours on Located",
         "Green = mailing address. Blue = AI. Gold = hand."),
    ])
    _write_sheet(wb, "Located", located, freeze_cols=1, stripe=True,
                 wrap={"institution"}, source_col="source")
    if not blank.empty:
        show = blank.reindex(columns=["institution", "n_events", "notes"])
        _write_sheet(wb, "Not located", show, freeze_cols=1, stripe=True,
                     wrap={"institution", "notes"})
    wb.close()


def build_codebook(stats: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    section = Font(bold=True, size=13, color="1F3864")
    key_font = Font(bold=True, color="1F3864")

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 100
    ws.sheet_properties.tabColor = "1F3864"

    lines = [
        ("Dataset", "American Men of Science, 4th edition (1927) — scientist-year panel"),
        ("Source", "J. McKeen Cattell (ed.), American Men of Science: A Biographical "
                   "Directory, 4th ed., The Science Press, 1927."),
        ("Construction", "Biographical entries transcribed from page scans with a vision "
                         "LLM under a strict JSON schema, cross-validated against the "
                         "volume's own OCR text layer (entry rosters, name spellings, "
                         "birth data, star counts). Unverifiable values are removed "
                         "rather than guessed. Career years are never interpolated."),
        ("Locations", "Institution geography is a separate bridge file. "
                      "'mailing address' = the book's own printed 1927 address names "
                      "the institution. 'AI' = classify-then-locate draft, not yet "
                      "hand-verified. 'hand' = typed in by a person. Dual-city strings "
                      "and non-places are left blank on purpose — never geocoded from "
                      "the modern web."),
        ("Panel rule", "Activity is filled only for years the directory itself confirms "
                       "(a dated degree or a dated spell); gaps are left blank. "
                       "Concurrent roles occupy separate numbered slots."),
        ("Order", "Every table is sorted as the book is read: PDF page, then surname, "
                  "then year / start year. Excel files freeze Page + Scientist so "
                  "scrolling down a career feels like turning the page."),
        ("", ""),
        ("Scientists", "{:,}".format(stats["n_scientists"])),
        ("Panel rows (scientist-years)", "{:,}".format(stats["n_panel"])),
        ("Dated career events", "{:,}".format(stats["n_events"])),
        ("Starred (top-1000) scientists", "{:,}".format(stats["n_starred"])),
        ("Institutions located", "{:,} of {:,} distinct strings ({})".format(
            stats["n_inst_coded"], stats["n_inst"], stats["inst_coverage"])),
        ("  of which mailing address", "{:,} strings / {:,} events".format(
            stats["n_mail_rows"], stats["n_mail_events"])),
        ("  of which AI draft", "{:,} strings / {:,} events".format(
            stats["n_ai_rows"], stats["n_ai_events"])),
        ("  of which hand-coded", "{:,} strings / {:,} events".format(
            stats["n_hand_rows"], stats["n_hand_events"])),
        ("Built", dt.date.today().isoformat()),
        ("", ""),
        ("Files (Excel + CSV)", ""),
        (SUMMARY_OUT + ".xlsx / .csv", "One row per scientist, book order."),
        (EVENTS_OUT + ".xlsx / .csv", "One row per dated degree/position; locations merged."),
        (PANEL_OUT + ".xlsx / .csv", "Balanced scientist-year panel (main analysis file)."),
        (LOCATIONS_OUT + ".xlsx / .csv", "Institution → city/state/country, with source."),
        (CODEBOOK, "This workbook: overview + column dictionary."),
    ]
    for r, (k, v) in enumerate(lines, start=1):
        a = ws.cell(row=r, column=1, value=k)
        b = ws.cell(row=r, column=2, value=v)
        a.font = key_font
        b.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 36 if len(str(v)) > 90 else 18
    ws["A1"].font = ws["A21"].font = section
    ws.freeze_panes = "A2"

    def sheet(title, cols):
        w = wb.create_sheet(title)
        for c, h in enumerate(["Column", "Description"], start=1):
            cell = w.cell(row=1, column=c, value=h)
            cell.fill, cell.font = header_fill, header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        w.column_dimensions["A"].width = 44
        w.column_dimensions["B"].width = 100
        w.row_dimensions[1].height = 22
        for r, (col, desc) in enumerate(cols, start=2):
            w.cell(row=r, column=1, value=col).font = Font(bold=True)
            w.cell(row=r, column=2, value=desc).alignment = Alignment(
                wrap_text=True, vertical="top")
            w.row_dimensions[r].height = 32
        w.freeze_panes = "A2"
        w.sheet_properties.tabColor = "1F3864"
        w.auto_filter.ref = "A1:B%d" % (len(cols) + 1)
        _ = get_column_letter

    sheet("Panel columns", PANEL_COLS)
    sheet("Events columns", EVENTS_COLS)
    sheet("Scientists columns", SUMMARY_COLS)
    sheet("Locations columns", LOCATION_COLS)
    wb.save(RELEASE / CODEBOOK)


def _location_stats(full: pd.DataFrame, located: pd.DataFrame) -> dict:
    n_events = pd.to_numeric(full["n_events"], errors="coerce").fillna(0)
    located = located.copy()
    located["n_events"] = pd.to_numeric(located["n_events"], errors="coerce").fillna(0)
    by = located.groupby("source", dropna=False)["n_events"]
    counts = located.groupby("source", dropna=False).size()

    def rows(src):
        return int(counts.get(src, 0))

    def events(src):
        return int(by.sum().get(src, 0))

    has_geo = full[["city", "state_region", "country"]].replace("", pd.NA).notna().any(axis=1)
    with_inst_events = int(n_events.sum())
    located_events = int(located["n_events"].sum())
    coverage = "%.1f%% of events" % (100 * located_events / max(with_inst_events, 1))
    return {
        "n_inst": len(full),
        "n_inst_coded": int(has_geo.sum()),
        "n_blank": int((~has_geo).sum()),
        "n_mail_rows": rows("mailing address"),
        "n_mail_events": events("mailing address"),
        "n_ai_rows": rows("AI"),
        "n_ai_events": events("AI"),
        "n_hand_rows": rows("hand"),
        "n_hand_events": events("hand"),
        "inst_coverage": coverage,
        "located_events": located_events,
        "all_events_in_bridge": with_inst_events,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--no-excel", action="store_true",
                        help="Write CSVs + codebook only (skip the large Excel files).")
    parser.add_argument("--no-panel-xlsx", action="store_true",
                        help="Skip only the ~600k-row panel workbook.")
    args = parser.parse_args()

    for name in (PANEL_IN, EVENTS_IN, SUMMARY_IN):
        if not (BASE / name).exists():
            raise SystemExit("Missing %s -- run 'python extract_panel.py "
                             "--panel-only' first." % name)
    RELEASE.mkdir(exist_ok=True)

    print("Loading working tables ...")
    panel = pd.read_csv(BASE / PANEL_IN, dtype=str, low_memory=False)
    events = pd.read_csv(BASE / EVENTS_IN, dtype=str, low_memory=False)
    summary = pd.read_csv(BASE / SUMMARY_IN, dtype=str, low_memory=False)

    leftover = [c for c in events.columns if c in
                ("n_events", "notes", "institution_city", "institution_state_region",
                 "institution_country", "institution_location_source")]
    if leftover:
        events = events.drop(columns=leftover)

    loc_path = BASE / LOCATIONS_IN
    loc_stats = {
        "n_inst": 0, "n_inst_coded": 0, "n_blank": 0,
        "n_mail_rows": 0, "n_mail_events": 0,
        "n_ai_rows": 0, "n_ai_events": 0,
        "n_hand_rows": 0, "n_hand_events": 0,
        "inst_coverage": "0%",
        "located_events": 0, "all_events_in_bridge": 0,
    }
    located = pd.DataFrame(columns=LOCATION_ORDER)
    blank = pd.DataFrame()

    if loc_path.exists():
        full = pd.read_csv(loc_path, dtype=str).fillna("")
        full["source"] = full["notes"].map(source_from_notes) if "notes" in full.columns else "hand"
        has_geo = full[["city", "state_region", "country"]].replace("", pd.NA).notna().any(axis=1)
        located = full.loc[has_geo, LOCATION_ORDER].copy()
        located["n_events"] = pd.to_numeric(located["n_events"], errors="coerce")
        located = located.sort_values(["n_events", "institution"],
                                      ascending=[False, True]).reset_index(drop=True)
        blank = full.loc[~has_geo].copy()
        blank["n_events"] = pd.to_numeric(blank["n_events"], errors="coerce")
        blank = blank.sort_values(["n_events", "institution"],
                                  ascending=[False, True]).reset_index(drop=True)
        loc_stats = _location_stats(full, located)

        lookup = load_location_lookup(loc_path)
        bridge = lookup.rename(columns={
            "institution": "institution_organization",
            "city": "institution_city",
            "state_region": "institution_state_region",
            "country": "institution_country",
            "source": "institution_location_source",
        })[["institution_organization", "institution_city",
            "institution_state_region", "institution_country",
            "institution_location_source"]]
        events = events.merge(bridge, on="institution_organization", how="left")
        located.to_csv(RELEASE / (LOCATIONS_OUT + ".csv"), index=False, encoding="utf-8-sig")

    summary = _book_sort(_reorder(summary, SUMMARY_ORDER), [])
    events = _book_sort(_reorder(events, EVENTS_ORDER), ["start_year"])
    panel = _book_sort(_reorder(panel, PANEL_FRONT), ["year"])

    for frame, ints in (
        (summary, ["source_pdf_page", "birth_year", "star_status",
                   "n_degrees", "n_study_spells", "n_positions", "n_parallel_positions"]),
        (events, ["source_pdf_page", "start_year", "end_year", "birth_year",
                  "star_status", "is_current_1927_role"]),
        (panel, ["source_pdf_page", "year", "age", "birth_year",
                 "star_status", "activity_confirmed", "is_current_1927_role"]),
    ):
        for col in ints:
            if col in frame.columns:
                frame[col] = pd.to_numeric(frame[col], errors="coerce")

    print("Writing CSVs ...")
    panel.to_csv(RELEASE / (PANEL_OUT + ".csv"), index=False, encoding="utf-8-sig")
    events.to_csv(RELEASE / (EVENTS_OUT + ".csv"), index=False, encoding="utf-8-sig")
    summary.to_csv(RELEASE / (SUMMARY_OUT + ".csv"), index=False, encoding="utf-8-sig")

    stats = {
        "n_scientists": len(summary),
        "n_panel": len(panel),
        "n_events": len(events),
        "n_starred": int(pd.to_numeric(summary.get("star_status"),
                                       errors="coerce").fillna(0).sum()),
        **loc_stats,
    }
    build_codebook(stats)

    if not args.no_excel:
        write_scientists_xlsx(summary, stats)
        if loc_path.exists():
            write_locations_xlsx(located, blank, stats)
        write_events_xlsx(events, stats)
        if not args.no_panel_xlsx:
            write_panel_xlsx(panel, stats)

    print("Release written to %s:" % RELEASE.name)
    for f in sorted(RELEASE.iterdir()):
        print("  %-42s %8.1f MB" % (f.name, f.stat().st_size / 1_048_576))


if __name__ == "__main__":
    main()
